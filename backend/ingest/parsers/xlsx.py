"""엑셀(xlsx/xlsb) 파서 — 모델·리포트 예시 인제스트 + 복호화.

두 특수과제:
  ① 살아있는 수식 vs 캐시값: data_only=True 면 캐시값(엔진투입), 수식 문자열은 provenance
     에 보존(감사인이 '=C4/C5' 추적).
  ② 복호화 + ㅁ→a: 암호화 파일(pe양식)은 msoffcrypto 로 복호화. 파일명에 두벌식 자모로
     비번 인코딩(`1ㅁ2ㅁ3ㅁ` → ㅁ=물리 'a'키 → `1a2a3a`)된 관행을 자동 복원.

openpyxl·msoffcrypto 는 선택 의존성(lazy import). 미설치 시 명확한 오류.
셀 출처 = (sheet, cell) locator. char span 없음(엑셀은 셀 좌표가 곧 위치).
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass

from ..provenance import ExtractMethod, Locator, SourceKind
from .base import BaseParser, ParseResult

# ── 두벌식 자모 → QWERTY 물리키 (파일명 비번 복원) ────────────────────────────
_DUBEOLSIK = {
    "ㅂ": "q", "ㅈ": "w", "ㄷ": "e", "ㄱ": "r", "ㅅ": "t", "ㅛ": "y", "ㅕ": "u",
    "ㅑ": "i", "ㅐ": "o", "ㅔ": "p", "ㅁ": "a", "ㄴ": "s", "ㅇ": "d", "ㄹ": "f",
    "ㅎ": "g", "ㅗ": "h", "ㅓ": "j", "ㅏ": "k", "ㅣ": "l", "ㅋ": "z", "ㅌ": "x",
    "ㅊ": "c", "ㅍ": "v", "ㅠ": "b", "ㅜ": "n", "ㅡ": "m",
}


def jamo_to_qwerty(s: str) -> str:
    """두벌식 자모를 같은 물리키 영문으로 치환(그 외 문자는 그대로)."""
    return "".join(_DUBEOLSIK.get(ch, ch) for ch in s)


def password_from_stem(stem: str) -> str | None:
    """파일명 어간에서 비번 후보 추출: 마지막 '_' 뒤 토큰의 자모를 영문키로 치환.

    예: 'DCF_클래시스_1ㅁ2ㅁ3ㅁ' → '1a2a3a'. '_' 없거나 순수 한글이면 None.
    """
    if "_" not in stem:
        return None
    tail = stem.rsplit("_", 1)[1]
    pw = jamo_to_qwerty(tail)
    # 영문/숫자/기호로 구성된 그럴싸한 비번만
    if pw and re.fullmatch(r"[A-Za-z0-9!@#$%^&*]+", pw):
        return pw
    return None


def _load_workbook(path: str, *, password: str | None, data_only: bool):
    """openpyxl 워크북 로드(암호화면 msoffcrypto 복호화 후)."""
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl 필요: pip install openpyxl") from e

    if password is None:
        try:
            return openpyxl.load_workbook(path, read_only=True, data_only=data_only)
        except Exception:
            # 암호화 가능성 → 파일명 비번 자동시도
            from pathlib import Path
            guess = password_from_stem(Path(path).stem)
            if guess is None:
                raise
            password = guess

    try:
        import msoffcrypto
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("암호화 파일: pip install msoffcrypto-tool") from e
    buf = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(buf)
    import openpyxl
    return openpyxl.load_workbook(buf, read_only=True, data_only=data_only)


@dataclass
class SheetCell:
    sheet: str
    coord: str            # 'C44'
    value: object
    formula: str | None   # data_only 로는 못 얻음; 필요 시 별도 로드


class XlsxParser(BaseParser):
    """엑셀 → 비어있지 않은 셀을 ProvenancedValue 로 방출.

    data_only=True(기본): 캐시값 방출. capture_formulas=True 면 수식도 별도 로드해
    provenance.note 에 부착(감사추적). password 미지정 시 파일명 두벌식 비번 자동시도.
    """
    source_kind = SourceKind.XLSX
    default_method = ExtractMethod.FORMULA

    def __init__(self, source_id: str, *, password: str | None = None,
                 sheets: list[str] | None = None,
                 capture_formulas: bool = False) -> None:
        super().__init__(source_id)
        self.password = password
        self.only_sheets = sheets
        self.capture_formulas = capture_formulas
        self.sheet_names: list[str] = []

    def extract(self, raw: object) -> ParseResult:
        """raw = xlsx 경로."""
        path = str(raw)
        wb = _load_workbook(path, password=self.password, data_only=True)
        self.sheet_names = list(wb.sheetnames)
        formulas = self._load_formulas(path) if self.capture_formulas else {}

        for sname in wb.sheetnames:
            if self.only_sheets and sname not in self.only_sheets:
                continue
            ws = wb[sname]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is None or (isinstance(v, str) and v.strip() == ""):
                        continue
                    coord = cell.coordinate
                    note = None
                    if self.capture_formulas:
                        f = formulas.get((sname, coord))
                        if f:
                            note = f"formula={f}"
                    self.emit(
                        f"{sname}!{coord}", v,
                        locator=Locator(sheet=sname, cell=coord),
                        note=note,
                    )
        return self.result

    def _load_formulas(self, path: str) -> dict[tuple[str, str], str]:
        """data_only=False 로 재로드해 수식 문자열 수집(=로 시작하는 셀)."""
        wb = _load_workbook(path, password=self.password, data_only=False)
        out: dict[tuple[str, str], str] = {}
        for sname in wb.sheetnames:
            if self.only_sheets and sname not in self.only_sheets:
                continue
            for row in wb[sname].iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        out[(sname, cell.coordinate)] = cell.value
        return out
