"""외부 xlsx 파서(openpyxl 기반) 테스트 — 셀 방출·비번 헬퍼.

우리 DCF 모델 왕복은 test_dcf_roundtrip(stdlib). 여기선 외부 임의 xlsx(리포트 예시·
암호화 모델) 인제스트용 XlsxParser. openpyxl 없으면 파서 테스트는 skip.
stdlib: `python tests/test_xlsx_parser.py`
"""
from __future__ import annotations

import sys
import tempfile
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from ingest.parsers.xlsx import (  # noqa: E402
    XlsxParser, jamo_to_qwerty, password_from_stem,
)

try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── 비번 헬퍼(두벌식 자모 → 물리키) — 순수함수, 의존 없음 ─────────────────────
def test_jamo_to_qwerty():
    assert jamo_to_qwerty("ㅁ") == "a"
    assert jamo_to_qwerty("1ㅁ2ㅁ3ㅁ") == "1a2a3a"
    assert jamo_to_qwerty("ㅂㅈㄷ") == "qwe"


def test_password_from_stem():
    assert password_from_stem("DCF_클래시스_1ㅁ2ㅁ3ㅁ") == "1a2a3a"
    assert password_from_stem("Simplified_DCF_솔루엠_1ㅁ2ㅁ3ㅁ") == "1a2a3a"
    assert password_from_stem("이름없음") is None          # '_' 없음
    assert password_from_stem("파일_순수한글") is None       # 자모치환해도 비번꼴 아님


# ── XlsxParser 셀 방출 (openpyxl 필요) ────────────────────────────────────────
def _make_xlsx() -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "리포트"
    ws["B2"] = "WACC"; ws["C2"] = 0.0624
    ws["B3"] = "매출"; ws["C3"] = 227400
    ws["B4"] = ""      # 빈 문자열 → 스킵
    p = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    wb.save(p)
    return p


def test_xlsx_parser_emits_cells():
    if not HAS_OPENPYXL:
        print("  (skip: openpyxl 없음)"); return
    p = XlsxParser("리포트예시")
    res = p.extract(_make_xlsx())
    assert res.value_of("리포트!C2") == Decimal("0.0624")
    assert res.value_of("리포트!C3") == Decimal("227400")
    v = res.by_name("리포트!C2")
    assert v.provenance.locator.sheet == "리포트" and v.provenance.locator.cell == "C2"
    assert v.provenance.source_kind.value == "xlsx"
    # 빈 셀은 방출 안 함
    assert res.by_name("리포트!B4") is None


def test_xlsx_parser_sheet_filter():
    if not HAS_OPENPYXL:
        print("  (skip)"); return
    p = XlsxParser("x", sheets=["없는시트"])
    res = p.extract(_make_xlsx())
    assert len(res.values) == 0                    # 대상 시트 없음 → 방출 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass
    import traceback
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    ok = 0
    for fn in fns:
        try:
            fn(); ok += 1; print(f"  ok  {fn.__name__}")
        except Exception:
            print(f"  FAIL {fn.__name__}"); traceback.print_exc()
    print(f"\n{ok}/{len(fns)} passed")
